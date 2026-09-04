"""Enterprise-styled report exports (CSV / XLSX / PDF) with executive summary.

PDF: page 1 = executive summary (KPI grid, key findings, recommendations),
then a styled detail table with per-page footer. XLSX: summary sheet + styled
detail sheet (frozen header, autofilter, status colors). CSV stays machine-
readable (no styling) with the same columns.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app_v4.core.utcdatetime import utc_now

# Brand palette (matches the ops-terminal UI: dark ink + amber accent).
_INK = colors.HexColor("#1a1a1a")
_AMBER = colors.HexColor("#b8860b")  # darker amber, readable on white
_GREEN = colors.HexColor("#1e8449")
_RED = colors.HexColor("#c0392b")
_ZEBRA = colors.HexColor("#f2f2f2")
_HEADER_FG = colors.whitesmoke
_XL_HEADER_BG = "1A1A1A"
_XL_HEADER_FG = "F5F5F5"


@dataclass(frozen=True)
class BackupReportRow:
    id: int
    switch_name: str
    taken_at: datetime
    backup_type: str
    success: bool
    size_bytes: int
    message: str


@dataclass(frozen=True)
class ComplianceRow:
    """One switch's config-management status (ISO 27001 A.8.9 evidence)."""

    switch: str
    ip: str
    model: str
    baseline: str  # "yes" | "no" | "stale"
    last_backup: str
    open_reviews: int
    last_review: str
    review_state: str  # pending | in_review | approved | flagged | dismissed | ""
    next_review: str = ""  # YYYY-MM-DD when the reminder-review cycle comes due


HEADERS = ["ID", "Switch", "Taken at", "Type", "Status", "Size (KB)", "Message"]
COMPLIANCE_HEADERS = [
    "Switch", "IP", "Model", "Baseline", "Last backup",
    "Open reviews", "Last review", "Review state", "Next review",
]


def _row_values(row: BackupReportRow) -> list[str]:
    return [
        str(row.id),
        row.switch_name,
        row.taken_at.strftime("%Y-%m-%d %H:%M:%S"),
        row.backup_type,
        "ok" if row.success else "failed",
        f"{round((row.size_bytes or 0) / 1024, 1)}",
        (row.message or "").replace("\n", " ").strip(),
    ]


def _compliance_values(row: ComplianceRow) -> list[str]:
    return [
        row.switch,
        row.ip,
        row.model or "",
        row.baseline,
        row.last_backup,
        str(row.open_reviews),
        row.last_review,
        row.review_state,
        row.next_review,
    ]


# ---------------------------------------------------------------------------
# Executive summary: KPIs, findings, recommendations
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ComplianceKPI:
    total_switches: int = 0
    covered: int = 0
    missing: list[str] = field(default_factory=list)
    stale: list[str] = field(default_factory=list)
    pending: int = 0
    flagged: int = 0
    approved: int = 0
    failed_backups: int = 0
    total_backups: int = 0
    coverage_pct: int = 0
    generated_at: str = ""


def compliance_kpis(rows: list[ComplianceRow]) -> ComplianceKPI:
    total = len(rows)
    missing = [r.switch for r in rows if r.baseline == "no"]
    stale = [r.switch for r in rows if r.baseline == "stale"]
    covered = total - len(missing)
    return ComplianceKPI(
        total_switches=total,
        covered=covered,
        missing=missing,
        stale=stale,
        pending=sum(r.open_reviews for r in rows),
        flagged=sum(1 for r in rows if r.review_state == "flagged"),
        approved=sum(1 for r in rows if r.review_state == "approved"),
        coverage_pct=round(100 * covered / total) if total else 0,
        generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S UTC"),
    )


def backup_kpis(rows: list[BackupReportRow]) -> ComplianceKPI:
    return ComplianceKPI(
        total_backups=len(rows),
        failed_backups=sum(1 for r in rows if not r.success),
        pending=sum(1 for r in rows if r.success and "Perubahan konfigurasi terdeteksi" in (r.message or "")),
        generated_at=utc_now().strftime("%Y-%m-%d %H:%M:%S UTC"),
    )


def _findings(k: ComplianceKPI) -> list[str]:
    out: list[str] = []
    if k.missing:
        out.append(
            f"{len(k.missing)} switch tanpa baseline ({', '.join(k.missing[:6])}"
            + ("…" if len(k.missing) > 6 else "")
            + ") — drift tidak terdeteksi pada switch tersebut."
        )
    if k.stale:
        out.append(
            f"{len(k.stale)} baseline melewati jadwal reminder review ({', '.join(k.stale[:6])}"
            + ("…" if len(k.stale) > 6 else "")
            + ")."
        )
    if k.pending:
        out.append(f"{k.pending} review menunggu keputusan.")
    if k.flagged:
        out.append(f"{k.flagged} review berstatus FLAGGED — perlu tindak lanjut.")
    if k.total_backups and k.failed_backups:
        out.append(
            f"{k.failed_backups} dari {k.total_backups} backup gagal pada periode laporan."
        )
    if not out:
        out.append("Tidak ada temuan yang memerlukan tindakan korektif pada periode ini.")
    return out


def _recommendations(k: ComplianceKPI) -> list[str]:
    recs: list[str] = []
    if k.missing:
        recs.append("Buat baseline untuk switch yang belum ter-cover (halaman Baselines) agar drift detection aktif.")
    if k.stale:
        recs.append("Jalankan tombol Review pada baseline yang due untuk re-attestation dan reset siklus.")
    if k.pending:
        recs.append("Selesaikan review pending di halaman Config Review — setiap drift wajib memiliki keputusan terdokumentasi.")
    if k.failed_backups:
        recs.append("Investigasi backup yang gagal (halaman History) — periksa kredensial dan reachability switch.")
    if not recs:
        recs.append("Pertahankan siklus review saat ini; tidak ada tindakan korektif.")
    return recs


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------


def _kpi_pdf_table(k: ComplianceKPI, compliance_mode: bool) -> Table:
    styles = getSampleStyleSheet()
    if compliance_mode:
        header = ["Switch", "Coverage", "Pending review", "Flagged", "Baseline due"]
        values = [str(k.total_switches), f"{k.coverage_pct}%", str(k.pending), str(k.flagged), str(len(k.stale))]
        color_rules = [
            ("TEXTCOLOR", (2, 1), (2, 1), _AMBER if k.pending else _GREEN),
            ("TEXTCOLOR", (3, 1), (3, 1), _RED if k.flagged else _GREEN),
            ("TEXTCOLOR", (4, 1), (4, 1), _RED if k.stale else _GREEN),
            ("TEXTCOLOR", (1, 1), (1, 1), _AMBER if k.coverage_pct < 100 else _GREEN),
        ]
    else:
        header = ["Total backup", "Gagal", "Perubahan config", "Success rate"]
        total = k.total_backups or 1
        values = [
            str(k.total_backups),
            str(k.failed_backups),
            str(k.pending),
            f"{round(100 * (k.total_backups - k.failed_backups) / total)}%",
        ]
        color_rules = [
            ("TEXTCOLOR", (1, 1), (1, 1), _RED if k.failed_backups else _GREEN),
            ("TEXTCOLOR", (3, 1), (3, 1), _GREEN if not k.failed_backups else _AMBER),
        ]
    table = Table([header, values], colWidths=[110, 110, 130, 110, 110][: len(header)])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _INK),
                ("TEXTCOLOR", (0, 0), (-1, 0), _HEADER_FG),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 1), (-1, 1), "CENTER"),
                *color_rules,
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return table


def _exec_summary_elements(k: ComplianceKPI, title: str, compliance_mode: bool) -> list:
    styles = getSampleStyleSheet()
    return [
        Paragraph("Executive Summary", styles["Title"]),
        Paragraph(f"<i>{title}</i>", styles["BodyText"]),
        Paragraph(f"Generated {k.generated_at}", styles["BodyText"]),
        Spacer(1, 14),
        _kpi_pdf_table(k, compliance_mode),
        Spacer(1, 14),
        Paragraph("<b>Key findings</b>", styles["Heading3"]),
        *[Paragraph(f"•  {f}", styles["BodyText"]) for f in _findings(k)],
        Spacer(1, 8),
        Paragraph("<b>Recommendations</b>", styles["Heading3"]),
        *[Paragraph(f"•  {r}", styles["BodyText"]) for r in _recommendations(k)],
        Spacer(1, 18),
    ]


def _page_footer(title: str):
    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(24, 14, f"{title} · generated {utc_now().strftime('%Y-%m-%d %H:%M UTC')} · Internal use")
        canvas.drawRightString(landscape(A4)[0] - 24, 14, f"Page {doc.page}")
        canvas.restoreState()

    return footer


def _style_detail_table(table: Table, status_col: int | None = None) -> None:
    """Dark header, zebra rows, status-column coloring."""
    style: list = [
        ("BACKGROUND", (0, 0), (-1, 0), _INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), _HEADER_FG),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, _ZEBRA]),
    ]
    if status_col is not None:
        for i, row in enumerate(table._cellvalues[1:], start=1):
            value = str(row[status_col]).lower()
            if value in ("failed", "flagged", "stale", "no"):
                style.append(("TEXTCOLOR", (status_col, i), (status_col, i), _RED))
            elif value in ("ok", "yes", "approved"):
                style.append(("TEXTCOLOR", (status_col, i), (status_col, i), _GREEN))
    table.setStyle(TableStyle(style))


def _style_xlsx_detail(sheet, headers: list[str], status_col: int | None = None) -> None:
    """Frozen styled header, autofilter, column widths, status colors."""
    header_fill = PatternFill("solid", fgColor=_XL_HEADER_BG)
    header_font = Font(bold=True, color=_XL_HEADER_FG)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    red_font = Font(color="C0392B")
    green_font = Font(color="1E8449")
    if status_col is not None:
        col_letter = get_column_letter(status_col + 1)
        for row_idx in range(2, sheet.max_row + 1):
            value = str(sheet[f"{col_letter}{row_idx}"].value or "").lower()
            if value in ("failed", "flagged", "stale", "no"):
                sheet[f"{col_letter}{row_idx}"].font = red_font
            elif value in ("ok", "yes", "approved"):
                sheet[f"{col_letter}{row_idx}"].font = green_font
    widths = {"A": 8, "B": 18, "C": 20, "D": 14, "E": 12, "F": 12, "G": 60, "H": 16, "I": 14, "J": 14}
    for idx in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(idx)].width = widths.get(get_column_letter(idx), 18)


def _add_xlsx_summary_sheet(wb, k: ComplianceKPI, title: str, compliance_mode: bool) -> None:
    sheet = wb.create_sheet("Executive Summary", 0)
    bold = Font(bold=True, size=14)
    section = Font(bold=True, size=11)
    sheet["A1"] = "Executive Summary"
    sheet["A1"].font = bold
    sheet["A2"] = title
    sheet["A3"] = f"Generated {k.generated_at}"
    row = 5
    if compliance_mode:
        kpis = [
            ("Total switch", k.total_switches),
            ("Baseline coverage", f"{k.coverage_pct}%"),
            ("Review pending", k.pending),
            ("Review flagged", k.flagged),
            ("Baseline due (reminder review)", len(k.stale)),
            ("Switch tanpa baseline", ", ".join(k.missing) or "-"),
        ]
    else:
        kpis = [
            ("Total backup", k.total_backups),
            ("Backup gagal", k.failed_backups),
            ("Backup dengan perubahan config", k.pending),
        ]
    for label, value in kpis:
        sheet.cell(row=row, column=1, value=label).font = section
        sheet.cell(row=row, column=2, value=value)
        row += 1
    row += 1
    sheet.cell(row=row, column=1, value="Key findings").font = section
    row += 1
    for finding in _findings(k):
        sheet.cell(row=row, column=1, value=f"• {finding}")
        row += 1
    row += 1
    sheet.cell(row=row, column=1, value="Recommendations").font = section
    row += 1
    for rec in _recommendations(k):
        sheet.cell(row=row, column=1, value=f"• {rec}")
        row += 1
    sheet.column_dimensions["A"].width = 60
    sheet.column_dimensions["B"].width = 40


# ----- backup report -----


def render_csv(rows: list[BackupReportRow]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for row in rows:
        writer.writerow(_row_values(row))
    return buf.getvalue().encode("utf-8-sig")


def render_xlsx(rows: list[BackupReportRow], title: str = "Backup report") -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    _add_xlsx_summary_sheet(wb, backup_kpis(rows), title, compliance_mode=False)
    sheet = wb.active
    sheet.title = "Backups"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(_row_values(row))
    _style_xlsx_detail(sheet, HEADERS, status_col=4)
    output = BytesIO()
    wb.properties.title = title
    wb.save(output)
    return output.getvalue()


def render_pdf(rows: list[BackupReportRow], title: str = "Backup report") -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=30,
        title=title,
    )
    k = backup_kpis(rows)
    elements: list = _exec_summary_elements(k, title, compliance_mode=False)
    data = [HEADERS] + [_row_values(row) for row in rows]
    table = Table(data, repeatRows=1, colWidths=[40, 110, 120, 60, 50, 60, 260])
    _style_detail_table(table, status_col=4)
    elements.append(Paragraph("Detail", getSampleStyleSheet()["Heading2"]))
    elements.append(table)
    doc.build(elements, onFirstPage=_page_footer(title), onLaterPages=_page_footer(title))
    return output.getvalue()


# ----- config-management compliance report (ISO 27001 A.8.9) -----


def render_compliance_csv(rows: list[ComplianceRow]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(COMPLIANCE_HEADERS)
    for row in rows:
        writer.writerow(_compliance_values(row))
    return buf.getvalue().encode("utf-8-sig")


def render_compliance_xlsx(rows: list[ComplianceRow], title: str = "Config compliance report") -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    _add_xlsx_summary_sheet(wb, compliance_kpis(rows), title, compliance_mode=True)
    sheet = wb.active
    sheet.title = "Compliance"
    sheet.append(COMPLIANCE_HEADERS)
    for row in rows:
        sheet.append(_compliance_values(row))
    _style_xlsx_detail(sheet, COMPLIANCE_HEADERS, status_col=3)
    output = BytesIO()
    wb.properties.title = title
    wb.save(output)
    return output.getvalue()


def render_compliance_pdf(rows: list[ComplianceRow], title: str = "Config compliance report") -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=30,
        title=title,
    )
    k = compliance_kpis(rows)
    elements: list = _exec_summary_elements(k, title, compliance_mode=True)
    data = [COMPLIANCE_HEADERS] + [_compliance_values(row) for row in rows]
    table = Table(data, repeatRows=1, colWidths=[90, 70, 90, 60, 90, 60, 90, 90, 70])
    _style_detail_table(table, status_col=3)
    elements.append(Paragraph("Detail per switch", getSampleStyleSheet()["Heading2"]))
    elements.append(table)
    doc.build(elements, onFirstPage=_page_footer(title), onLaterPages=_page_footer(title))
    return output.getvalue()
